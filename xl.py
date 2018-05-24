#엑셀파일을 읽어오고 데이터를 뿌리는 연습을해보자
#pip3 install openpyxl을 설치해야한다.
import openpyxl
book = openpyxl.load_workbook("stat_104102.xlsx")
#1번째 방법
#print(book.get_sheet_names())
#print(book.get_sheet_by_name("stat_104102"))
#2번째 방법
sheet = book.worksheets[0]
for row in sheet.rows:
    for data in row:
        print(data.value, end=" ")#.value라고 해서 가져오는것이 중요하다
    print("", end="\n")


#데이터쓰기(엑셀)
workbook = openpyxl.Workbook()    
sheet = workbook.active
sheet["A1"]="테스트 파일"
sheet["A2"]="안녕하세요"
sheet.merge_cells("A1:C1")#A1부터 C1까지 병합된 모습을 볼수있다.
sheet["A1"].font = openpyxl.styles.Font(size=20,color="FF0000")
workbook.save("newfile.xlsx")